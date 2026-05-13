import os, re
import time
import random

import numpy as np
import calendar
import time
import csv
import pypandoc
import pandas as pd

import shutil
from enum import Enum
import subprocess
from pathlib import Path

import torch
import torchvision.transforms as transforms
from moviepy.editor import VideoFileClip
from openpyxl import Workbook, load_workbook


def set_random_seed(seed):
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    torch.cuda.manual_seed_all(seed)

class Summary(Enum):
    NONE = 0
    AVERAGE = 1
    SUM = 2
    COUNT = 3

class AverageMeter(object):
    """Computes and stores the average and current value"""
    def __init__(self, name, fmt=':f', summary_type=Summary.AVERAGE):
        self.name = name
        self.fmt = fmt
        self.summary_type = summary_type
        self.reset()

    def reset(self):
        self.val = 0
        self.avg = 0
        self.sum = 0
        self.count = 0

    def update(self, val, n=1):
        self.val = val
        self.sum += val * n
        self.count += n
        self.avg = self.sum / self.count

    def __str__(self):
        fmtstr = '{name} {val' + self.fmt + '} ({avg' + self.fmt + '})'
        return fmtstr.format(**self.__dict__)
    
    def summary(self):
        fmtstr = ''
        if self.summary_type is Summary.NONE:
            fmtstr = ''
        elif self.summary_type is Summary.AVERAGE:
            fmtstr = '{name} {avg:.3f}'
        elif self.summary_type is Summary.SUM:
            fmtstr = '{name} {sum:.3f}'
        elif self.summary_type is Summary.COUNT:
            fmtstr = '{name} {count:.3f}'
        else:
            raise ValueError('invalid summary type %r' % self.summary_type)
        
        return fmtstr.format(**self.__dict__)


class ProgressMeter(object):
    def __init__(self, num_batches, meters, prefix=""):
        self.batch_fmtstr = self._get_batch_fmtstr(num_batches)
        self.meters = meters
        self.prefix = prefix

    def display(self, batch):
        entries = [self.prefix + self.batch_fmtstr.format(batch)]
        entries += [str(meter) for meter in self.meters]
        print('\t'.join(entries))
        
    def display_summary(self):
        entries = [" *"]
        entries += [meter.summary() for meter in self.meters]
        print(' '.join(entries))

    def _get_batch_fmtstr(self, num_batches):
        num_digits = len(str(num_batches // 1))
        fmt = '{:' + str(num_digits) + 'd}'
        return '[' + fmt + '/' + fmt.format(num_batches) + ']'


def accuracy(output, target, topk=(1,)):
    """Computes the accuracy over the k top predictions for the specified values of k"""
    with torch.no_grad():
        maxk = max(topk)
        batch_size = target.size(0)

        _, pred = output.topk(maxk, 1, True, True)
        pred = pred.t()
        correct = pred.eq(target.view(1, -1).expand_as(pred))

        res = []
        for k in topk:
            correct_k = correct[:k].reshape(-1).float().sum(0, keepdim=True)
            res.append(correct_k.mul_(100.0 / batch_size))
        return res
        

def load_model_weight(load_path, model, device, args):
    if os.path.isfile(load_path):
        print("=> loading checkpoint '{}'".format(load_path))
        checkpoint = torch.load(load_path, map_location=device)
        state_dict = checkpoint['state_dict']
        # Ignore fixed token vectors
        if "token_prefix" in state_dict:
            del state_dict["token_prefix"]

        if "token_suffix" in state_dict:
            del state_dict["token_suffix"]

        args.start_epoch = checkpoint['epoch']
        try:
            best_acc1 = checkpoint['best_acc1']
        except:
            best_acc1 = torch.tensor(0)
        if device != 'cpu':
            # best_acc1 may be from a checkpoint from a different GPU
            best_acc1 = best_acc1.to(device)
        try:
            model.load_state_dict(state_dict)
        except:
            # TODO: implement this method for the generator class
            model.prompt_generator.load_state_dict(state_dict, strict=False)
        print("=> loaded checkpoint '{}' (epoch {})"
              .format(load_path, checkpoint['epoch']))
        del checkpoint
        torch.cuda.empty_cache()
    else:
        print("=> no checkpoint found at '{}'".format(load_path))


def validate(val_loader, model, criterion, args, output_mask=None):
    batch_time = AverageMeter('Time', ':6.3f', Summary.NONE)
    losses = AverageMeter('Loss', ':.4e', Summary.NONE)
    top1 = AverageMeter('Acc@1', ':6.2f', Summary.AVERAGE)
    top5 = AverageMeter('Acc@5', ':6.2f', Summary.AVERAGE)
    progress = ProgressMeter(
        len(val_loader),
        [batch_time, losses, top1, top5],
        prefix='Test: ')

    # switch to evaluate mode
    model.eval()

    with torch.no_grad():
        end = time.time()
        for i, (images, target) in enumerate(val_loader):
            if args.gpu is not None:
                images = images.cuda(args.gpu, non_blocking=True)
            if torch.cuda.is_available():
                target = target.cuda(args.gpu, non_blocking=True)

            # compute output
            with torch.cuda.amp.autocast():
                output = model(images)
                if output_mask:
                    output = output[:, output_mask]
                loss = criterion(output, target)

            # measure accuracy and record loss
            acc1, acc5 = accuracy(output, target, topk=(1, 2))
            losses.update(loss.item(), images.size(0))
            top1.update(acc1[0], images.size(0))
            top5.update(acc5[0], images.size(0))

            # measure elapsed time
            batch_time.update(time.time() - end)
            end = time.time()

            if i % args.print_freq == 0:
                progress.display(i)
        progress.display_summary()

    return top1.avg

def create_target_folders(root_path, folder_name, target_name, timestamp, split_files=False):
    folder_path = os.path.join(root_path, folder_name)
    if not os.path.exists(folder_path):
        os.makedirs(folder_path, exist_ok=True)
    
    # dest_path = os.path.join(folder_path, config.ALL_SOURCES_FOLDER) if train_source else 
    dest_path = os.path.join(folder_path, str(target_mapping(target_name)) + "-" +target_name)
    if not os.path.exists(dest_path):
        os.makedirs(dest_path, exist_ok=True)
    target_files_path = os.path.join(dest_path, "splitfiles" if split_files else "files")#files, splitfiles
    if not os.path.exists(target_files_path):
        os.makedirs(target_files_path, exist_ok=True)

    if timestamp is None:
        timestamp = calendar.timegm(time.gmtime())
        target_timestamp_path = os.path.join(dest_path, str(timestamp))
        os.makedirs(target_timestamp_path, exist_ok=True)
    else:
        target_timestamp_path = os.path.join(dest_path, str(timestamp))
    target_weights_path = os.path.join(target_timestamp_path, "weights")
    if not os.path.exists(target_weights_path):
        os.makedirs(target_weights_path, exist_ok=True)
    
    return target_files_path, target_weights_path, str(timestamp)

def target_mapping(target_name):
    if target_name == "081014_w_27":
        return 1
    elif target_name == "101609_m_36":
        return 2
    elif target_name == "112009_w_43":
        return 3
    elif target_name == "091809_w_43":
        return 4
    elif target_name == "071309_w_21":
        return 5
    elif target_name == "073114_m_25":
        return 6
    elif target_name == "080314_w_25":
        return 7
    elif target_name == "073109_w_28":
        return 8
    elif target_name == "100909_w_65":
        return 9
    elif target_name == "081609_w_40":
        return 10
    else:
        return 0

class MetricLogger:
    def __init__(self, save_dir="metrics", filename="metrics_log.csv"):
        os.makedirs(save_dir, exist_ok=True)
        self.filepath = os.path.join(save_dir, filename)

        # Write header once
        with open(self.filepath, mode="w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["epoch", "model", "au_sim_mean", "au_sim_std", "logits_mean", "logits_std"])

    def log(self, epoch, model_name, au_sim, logits):
        """Log metrics for each epoch."""
        au_sim_mean = au_sim.mean().item()
        au_sim_std = au_sim.std().item()
        logits_mean = logits.mean().item()
        logits_std = logits.std().item()

        with open(self.filepath, mode="a", newline="") as f:
            writer = csv.writer(f)
            writer.writerow([epoch, model_name, au_sim_mean, au_sim_std, logits_mean, logits_std])

        print(f"[Metrics] {model_name} | Epoch {epoch:03d} | "
              f"AU μ={au_sim_mean:.4f}, σ={au_sim_std:.4f}, "
              f"Logit μ={logits_mean:.4f}, σ={logits_std:.4f}")

def load_bah_src_subs(file_path, topk=None, selected_sub_list=None, index2value=False):
    subject_list = None
    unique_ids = extract_unique_ids(file_path)
    if topk is None:
        subject_list = unique_ids
        # subject_list = ['82553', '82554', '82555', '82557', '82563','82564', '82565', '82565', # 7
        #                  # 15    
        #                 ] # 76
    # 071709_w_23, 101814_m_58,080609_w_27,102008_w_22, 112310_m_20,112809_w_23,102316_w_50,071814_w_23,102214_w_36 // ontsne
    if selected_sub_list is not None:
        if index2value:
            return [subject_list[i] for i in selected_sub_list]
        else:
            # return [sub for sub in subject_list if sub in selected_sub_list]
            return [subject_list.index(sub) for sub in selected_sub_list if sub in subject_list]
    return subject_list

def extract_unique_ids(file_path):
    # Read file content
    with open(file_path, 'r') as file:
        content = file.read()

    # Extract all IDs that come after 'Videos/' using regex
    ids = re.findall(r'Videos/(\d+)/', content)

    # Get unique IDs
    unique_ids = sorted(set(ids))

    # Print results
    print("Total unique ID count:", len(unique_ids))
    # print("Unique ID list:")
    # print(unique_ids)

    return unique_ids


def create_source_files():
    # This script reads a master source text file and generates four new source files
    # containing only lines that belong to the specified source subject sets.

    # === CHANGE THIS PATH to your input file ===
    input_file = "lab_srcs78_082208w45_081714m36_112610w60_101908m61_071709w23_082014w24_110810m62_080209w26_101916m40_110614m42_____only.txt"  # <-- replace with your actual file name

    # ---- Source sets (from previous message) ----

    source_67 = [
    '081714_m_36','112610_w_60','071709_w_23','110810_m_62','080209_w_26',
    '110614_m_42','101814_m_58','071313_m_41','102514_w_40','101114_w_37',
    '100509_w_43','082315_w_60','120614_w_61','101514_w_36','092813_w_24',
    '102309_m_61','081617_m_27','080609_w_27','111313_m_64','071614_m_20',
    '101309_m_48','071911_w_24','102316_w_50','100417_m_44','083013_w_47',
    '083009_w_42','080714_m_23','101809_m_59','082909_m_47','101209_w_61',
    '092014_m_56','072414_m_23','101015_w_43','112909_w_20','111609_m_65',
    '100117_w_36','111409_w_63','080709_m_24','072714_m_23','112914_w_51',
    '120514_w_56','083109_m_60','110909_m_29','091814_m_37','071814_w_23',
    '092509_w_51','112809_w_23','100214_m_50','102214_w_36','082714_m_22',
    '082109_m_53','092808_m_51','080309_m_29','102008_w_22','111914_w_63',
    '082809_m_26','072514_m_27','082814_w_46','072609_w_23','101216_m_40',
    '091914_m_46','100914_m_39','112209_m_51','092514_m_50','092009_m_54',
    '082414_m_64','080614_m_24'
    ]

    source_57 = [
    '081714_m_36','112610_w_60','071709_w_23','110810_m_62','080209_w_26',
    '110614_m_42','101814_m_58','071313_m_41','102514_w_40','101114_w_37',
    '100509_w_43','082315_w_60','120614_w_61','101514_w_36','092813_w_24',
    '102309_m_61','081617_m_27','080609_w_27','111313_m_64','071614_m_20',
    '101809_m_59','082909_m_47','092014_m_56','101015_w_43','112909_w_20',
    '100117_w_36','111409_w_63','072714_m_23','112914_w_51','120514_w_56',
    '083109_m_60','110909_m_29','091814_m_37','071814_w_23','092509_w_51',
    '112809_w_23','100214_m_50','102214_w_36','082109_m_53','092808_m_51',
    '080309_m_29','102008_w_22','111914_w_63','082809_m_26','072514_m_27',
    '082814_w_46','072609_w_23','101216_m_40','091914_m_46','100914_m_39',
    '112209_m_51','092514_m_50','092009_m_54','082414_m_64','080614_m_24'
    ]

    source_47 = [
    '081714_m_36','112610_w_60','071709_w_23','110810_m_62','080209_w_26',
    '110614_m_42','101814_m_58','071313_m_41','102514_w_40','101114_w_37',
    '100509_w_43','082315_w_60','120614_w_61','101514_w_36','092813_w_24',
    '102309_m_61','081617_m_27','080609_w_27','111313_m_64','071614_m_20',
    '101809_m_59','082909_m_47','092014_m_56','101015_w_43','112909_w_20',
    '100117_w_36','111409_w_63','072714_m_23','112914_w_51','120514_w_56',
    '083109_m_60','110909_m_29','091814_m_37','071814_w_23',
    '072514_m_27','082814_w_46','072609_w_23','101216_m_40',
    '091914_m_46','100914_m_39','112209_m_51','092514_m_50',
    '092009_m_54','082414_m_64','080614_m_24'
    ]

    source_37 = [
    '081714_m_36','112610_w_60','071709_w_23','110810_m_62','080209_w_26',
    '110614_m_42','101814_m_58','071313_m_41','102514_w_40','101114_w_37',
    '100509_w_43','082315_w_60','120614_w_61','101514_w_36','092813_w_24',
    '102309_m_61','081617_m_27','080609_w_27','111313_m_64','071614_m_20',
    '101809_m_59','082909_m_47','092014_m_56','101015_w_43','112909_w_20',
    '100117_w_36','111409_w_63','072714_m_23','112914_w_51',
    '120514_w_56','083109_m_60','110909_m_29','091814_m_37',
    '071814_w_23','080614_m_24'
    ]

    # ===============================
    # Function to filter lines
    # ===============================

    def filter_lines(subject_list, lines):
        filtered = []
        for line in lines:
            for sub in subject_list:
                if f"/{sub}/" in line:
                    filtered.append(line)
                    break
        return filtered

    # ===============================
    # Read master file
    # ===============================

    with open(input_file, "r") as f:
        all_lines = f.readlines()

    # ===============================
    # Generate new files
    # ===============================

    output_sets = {
        "source_67.txt": source_67,
        "source_57.txt": source_57,
        "source_47.txt": source_47,
        "source_37.txt": source_37,
    }

    for filename, subject_list in output_sets.items():
        filtered_lines = filter_lines(subject_list, all_lines)

        with open(filename, "w") as out:
            out.writelines(filtered_lines)

        print(f"{filename} created with {len(filtered_lines)} lines.")

    print("Done.")

def create_target_subject_files():

    # =========================
    # CONFIG: set your input file
    # =========================
    INPUT_FILE = "sub_two_labels.txt"   # <-- change to your file path
    OUT_ROOT   = "WeightFiles"         # output root folder

    # Target subjects (numbering starts at 11)
    TARGET_SUBJECTS = [
        '082208_w_45','101908_m_61','082014_w_24','101916_m_40','112016_m_25','100514_w_51','112310_m_20','092714_m_64',
        '102414_w_58','083114_w_55','101309_m_48','071911_w_24','102316_w_50','100417_m_44','080714_m_23','101209_w_61',
        '072414_m_23','111609_m_65','080709_m_24','082714_m_22','092509_w_51','112809_w_23','100214_m_50','102214_w_36',
        '082109_m_53','092808_m_51','080309_m_29','102008_w_22','111914_w_63','082809_m_26','072514_m_27','082814_w_46',
        '072609_w_23','101216_m_40','091914_m_46','100914_m_39','112209_m_51','092514_m_50','092009_m_54','082414_m_64'
    ]

    # Build index -> subject mapping: 11_subject, 12_subject, ...
    indexed_subjects = [(11 + i, sub) for i, sub in enumerate(TARGET_SUBJECTS)]

    # Pre-create subject folders and open file handles
    os.makedirs(OUT_ROOT, exist_ok=True)

    # Prepare output writers
    writers = {}
    for idx, sub in indexed_subjects:
        folder_name = f"{idx:02d}_{sub}"  # 11_..., 12_..., etc (2-digit)
        out_dir = os.path.join(OUT_ROOT, folder_name, "files")
        os.makedirs(out_dir, exist_ok=True)

        out_path = os.path.join(out_dir, f"{sub}.txt")
        writers[sub] = open(out_path, "w", encoding="utf-8")

    # Read input and route lines to the correct subject file
    with open(INPUT_FILE, "r", encoding="utf-8") as f:
        for line in f:
            # Fast match: find subject token in the path segment "/<subject>/"
            for sub in TARGET_SUBJECTS:
                if f"/{sub}/" in line:
                    writers[sub].write(line)
                    break

    # Close files
    for fh in writers.values():
        fh.close()

    print(f"Done. Created {len(TARGET_SUBJECTS)} subject folders under '{OUT_ROOT}/'.")


def save_best_metrics(war, uar, f1, filename="best_metrics.csv"):
    # Create the data dictionary
    new_data = {
        'WAR': [war],
        'UAR': [uar],
        'F1': [f1]
    }
    df = pd.DataFrame(new_data)
    
    # Check if file exists to determine if we need a header
    file_exists = os.path.isfile(filename)
    
    # Append the results to the CSV
    df.to_csv(filename, mode='a', index=False, header=not file_exists)
    print(f"Metrics saved to {filename}")


def update_subject_result_xlsx(technique, filename, bs, temp, subject_name, subject_code, war, uar, f1, is_last_subject=False):
    """
    Updates a single subject's results in the Excel log.
    Maintains the horizontal layout: Metrics in Col 1, Subjects in Col 2 onwards.
    """
    strategy_label = f"Technique={technique}"
    
    # 1. Load or Create Workbook
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "Training Logs"
    else:
        wb = load_workbook(filename)
        ws = wb.active

    # 2. Find the correct experiment block
    # Search for the header that matches the strategy, batch size, and temperature
    target_start_row = None
    for row in range(ws.max_row, 0, -1):
        if ws.cell(row=row, column=1).value == strategy_label:
            if ws.cell(row=row+1, column=1).value == f"bs={bs}" and \
               ws.cell(row=row+2, column=1).value == f"temp={temp}":
                target_start_row = row
                break
    
    # 3. If no block exists, create the header and labels
    if target_start_row is None:
        target_start_row = ws.max_row + (2 if ws.max_row > 1 else 0)
        ws.cell(row=target_start_row, column=1, value=strategy_label)
        ws.cell(row=target_start_row+1, column=1, value=f"bs={bs}")
        ws.cell(row=target_start_row+2, column=1, value=f"temp={temp}")
        ws.cell(row=target_start_row+5, column=1, value="WAR")
        ws.cell(row=target_start_row+6, column=1, value="UAR")
        ws.cell(row=target_start_row+7, column=1, value="F1")

    # 4. Find the next available column (starting from Column 2 / B)
    # We check the 'Subject' row (row + 3) to find the first empty cell
    current_col = 2
    while ws.cell(row=target_start_row + 3, column=current_col).value is not None:
        if ws.cell(row=target_start_row + 3, column=current_col).value == "Avg":
            # If we already have an Avg column, we insert before it or just append
            break
        current_col += 1

    # 5. Write the subject data
    ws.cell(row=target_start_row + 3, column=current_col, value=subject_name)
    ws.cell(row=target_start_row + 4, column=current_col, value=subject_code)
    ws.cell(row=target_start_row + 5, column=current_col, value=float(war))
    ws.cell(row=target_start_row + 6, column=current_col, value=float(uar))
    ws.cell(row=target_start_row + 7, column=current_col, value=float(f1))

    # 6. If it's the last subject, calculate the Averages
    if is_last_subject:
        avg_col = current_col + 1
        ws.cell(row=target_start_row + 3, column=avg_col, value="Avg")
        ws.cell(row=target_start_row + 4, column=avg_col, value="Avg")
        
        # Calculate for WAR (row + 5), UAR (row + 6), and F1 (row + 7)
        for row_offset in range(5, 8):
            scores = []
            for col in range(2, avg_col): # Get scores from all subject columns
                val = float(ws.cell(row=target_start_row + row_offset, column=col).value)
                if isinstance(val, (int, float)):
                    scores.append(val)
            
            if scores:
                avg_val = sum(scores) / len(scores)
                ws.cell(row=target_start_row + row_offset, column=avg_col, value=round(avg_val, 4))

    # Save and close
    wb.save(filename)

def calculate_aggregate_performance(filename="experiment_results.csv"):
    if not os.path.isfile(filename):
        print("No results file found.")
        return
    
    # Read all stored results
    results_df = pd.read_csv(filename)
    
    # Calculate Mean and Standard Deviation
    avg_metrics = results_df.mean()
    std_metrics = results_df.std()
    num_runs = len(results_df)

    print(f"--- Aggregate Results over {num_runs} runs ---")
    print(f"Average WAR: {avg_metrics['WAR']:.4f} ± {std_metrics['WAR']:.4f}")
    print(f"Average UAR: {avg_metrics['UAR']:.4f} ± {std_metrics['UAR']:.4f}")
    print(f"Average F1:  {avg_metrics['F1']:.4f} ± {std_metrics['F1']:.4f}")
    
    return avg_metrics


def get_video_metadata(video_path):
    """
    Extract FPS (video) and sampling rate (audio) from a video file.

    Args:
        video_path (str): Path to video file

    Returns:
        fps (float): Frames per second of video
        audio_sr (int or None): Audio sampling rate (Hz), None if no audio
    """

    clip = VideoFileClip(video_path)
    # waveform, audio_sr = torchaudio.load(os.path.join(self.img_dir, 'audio/j1u8/j1u8_Counting3.wav'))

    # ---- Video FPS ----
    fps = clip.fps

    # ---- Audio SR ----
    if clip.audio is not None:
        audio_sr = clip.audio.fps  # sampling rate
    else:
        audio_sr = None

    clip.close()

    return fps, audio_sr


def extract_audio_from_folder(video_dir, audio_dir, sample_rate=16000):
    video_root = Path(video_dir)
    audio_root = Path(audio_dir)
    
    # Extensions to look for
    valid_extensions = ('.mp4', '.avi', '.mov', '.mkv', '.wmv')

    for root, dirs, files in os.walk(video_root):
        for file in files:
            if file.lower().endswith(valid_extensions):
                # Setup paths
                video_path = Path(root) / file
                
                # relative_to preserves the subfolder name
                relative_path = video_path.relative_to(video_root)
                audio_path = audio_root / relative_path.with_suffix('.wav')

                # Create the subfolder if it doesn't exist
                audio_path.parent.mkdir(parents=True, exist_ok=True)

                # FFmpeg command: 
                # -vn (no video), -ac 1 (mono), -ar (sample rate)
                command = [
                    'ffmpeg', '-y', '-i', str(video_path),
                    '-vn', '-acodec', 'pcm_s16le', 
                    '-ar', str(sample_rate), '-ac', '1', 
                    str(audio_path)
                ]

                try:
                    subprocess.run(command, check=True, capture_output=True)
                    print(f"Done: {relative_path}")
                except subprocess.CalledProcessError as e:
                    print(f"Failed: {video_path}. Error: {e.stderr.decode()}")